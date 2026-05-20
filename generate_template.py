#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GL HW Gaming Lab — Test Report Template Generator
Usage: python generate_template.py <form_data.json> [output.xlsx]

Reads EE report template files and generates a single workbook
with sheets matching the requested test items.
"""

import sys
import json
import copy
import re
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── Configuration ────────────────────────────────────────────────────────────

TEMPLATE_DIR = Path(r"G:\HW Gaming\EE 原始報告資料")

TEMPLATE_FILES = {
    'acc':    TEMPLATE_DIR / 'ACC.xlsx',
    'tpr':    TEMPLATE_DIR / 'TPR.xlsx',
    'sow':    TEMPLATE_DIR / 'SOW Perf.xlsx',
    'hwinfo': TEMPLATE_DIR / 'HW info.xlsx',
}

# Form checkbox value → sheet name in template file
ACC_SHEET_MAP = {
    'Cinebench R23 30-loop': 'Cinebench R23 30-loop',
    '3DMark Benchmark':      '3Dmark Benchmark',
}

TPR_SHEET_MAP = {
    'battery': 'Battery life',
    'perf':    'Performance',
    'igpu':    'iGPU',
}

SOW_SHEET_NAME   = 'RTX5060_X4 Performance test'
HWINFO_SHEET_IDX = 0  # use active sheet


# ─── Utilities ────────────────────────────────────────────────────────────────

def safe_str(v):
    if v is None: return ''
    if isinstance(v, dict): return v.get('Title') or v.get('Value') or ''
    return str(v)

def split_csv(s):
    if not s: return []
    return [x.strip() for x in str(s).split(',') if x.strip()]

def parse_items(text):
    """Extract numbered items: '1. Item A\n2. Item B' → ['Item A', 'Item B']"""
    items = []
    for line in (text or '').split('\n'):
        m = re.match(r'^\d+\.\s+(.+)', line.strip())
        if m:
            items.append(m.group(1).strip())
    return items

def parse_tpr(text):
    result = {
        'battery': False, 'perf': False, 'igpu': False,
        'batFan': '', 'perfFan': '', 'igpuFan': '',
    }
    for line in (text or '').split('\n'):
        l = line.strip()
        m_item = re.match(r'^\d+\.\s+(.+)', l)
        if m_item:
            v = m_item.group(1).lower()
            if 'battery' in v:
                result['battery'] = True
            elif 'igpu' in v:
                result['igpu'] = True
            else:
                result['perf'] = True
            continue
        m = re.search(r'Battery life Fan modes?:\s*(.+)', l, re.I)
        if m: result['batFan'] = m.group(1).strip()
        m = re.search(r'Performance Fan modes?:\s*(.+)', l, re.I)
        if m: result['perfFan'] = m.group(1).strip()
        m = re.search(r'iGPU Fan modes?:\s*(.+)', l, re.I)
        if m: result['igpuFan'] = m.group(1).strip()
    return result

def parse_sku_list(text):
    """Parse '1. SKU-X CPU:i7 GPU:RTX | 2. ...' into list of dicts."""
    skus = []
    for part in (text or '').split('|'):
        part = part.strip()
        if not part: continue
        m = re.match(r'^\d+\.\s*(\S+)(.*)', part)
        if not m: continue
        rest = m.group(2)
        def extract(key):
            mx = re.search(key + r':([^\s|]+)', rest, re.I)
            return mx.group(1) if mx else ''
        skus.append({
            'name':  m.group(1),
            'cpu':   extract('CPU'),
            'gpu':   extract('GPU'),
            'panel': extract('Panel'),
            'ddr':   extract('DDR'),
        })
    return skus


# ─── Sheet Copying ────────────────────────────────────────────────────────────

def copy_worksheet(src_ws, dst_wb, new_title):
    """
    Copy a worksheet into dst_wb.
    Preserves: values, styles, merges, row/col dimensions, freeze panes.
    Note: Charts and images cannot be copied by openpyxl.
    """
    ws = dst_wb.create_sheet(title=new_title[:31])

    # Column dimensions
    for col, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col].width  = dim.width
        ws.column_dimensions[col].hidden = dim.hidden

    # Row dimensions
    for row_idx, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[row_idx].height = dim.height
        ws.row_dimensions[row_idx].hidden = dim.hidden

    # Merged cells (before writing values)
    for rng in src_ws.merged_cells.ranges:
        ws.merge_cells(str(rng))

    # Cells
    for row in src_ws.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font        = copy.copy(cell.font)
                nc.fill        = copy.copy(cell.fill)
                nc.border      = copy.copy(cell.border)
                nc.alignment   = copy.copy(cell.alignment)
                nc.number_format = cell.number_format

    # Freeze panes
    if src_ws.freeze_panes:
        ws.freeze_panes = src_ws.freeze_panes

    # Tab color
    if src_ws.sheet_properties.tabColor:
        ws.sheet_properties.tabColor = copy.copy(src_ws.sheet_properties.tabColor)

    # Sheet format defaults
    if src_ws.sheet_format.defaultRowHeight:
        ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight

    return ws


# ─── Cover Sheet ──────────────────────────────────────────────────────────────

def build_cover_sheet(wb, form, idx=0):
    ws = wb.create_sheet(title='📋 需求資訊', index=idx)

    # Style helpers
    thin_side = Side(style='thin', color='9DC3E6')
    bdr = Border(top=thin_side, left=thin_side, bottom=thin_side, right=thin_side)
    c_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    r_align  = Alignment(horizontal='right',  vertical='center')

    hdr_font = Font(name='微軟正黑體', size=14, bold=True,  color='FFFFFFFF')
    cat_font = Font(name='微軟正黑體', size=12, bold=True,  color='FFFFFFFF')
    lbl_font = Font(name='微軟正黑體', size=11, bold=True,  color='FF1F3864')
    val_font = Font(name='微軟正黑體', size=11, bold=False, color='FF000000')
    tbl_font = Font(name='微軟正黑體', size=10, bold=True,  color='FFFFFFFF')
    dat_font = Font(name='微軟正黑體', size=10, bold=False, color='FF000000')

    hdr_fill = PatternFill('solid', fgColor='1F3864')
    cat_fill = PatternFill('solid', fgColor='2E74B5')
    lbl_fill = PatternFill('solid', fgColor='BDD7EE')
    val_fill = PatternFill('solid', fgColor='FFFDE7')
    row_even = PatternFill('solid', fgColor='DEEAF1')
    row_odd  = PatternFill('solid', fgColor='FFFFFF')

    # Column widths
    widths = [3, 20, 28, 20, 28, 3]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    r = 1

    def cell_styled(row, col, value, font, fill, align):
        c = ws.cell(row=row, column=col)
        c.value = value; c.font = font; c.fill = fill
        c.alignment = align; c.border = bdr
        return c

    def hrow(row, col1, col2, value, font, fill, align):
        ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
        return cell_styled(row, col1, value, font, fill, align)

    def lv(row, label, val, lc=2, vc=3):
        cell_styled(row, 1,  '',    lbl_font, lbl_fill, c_align)
        cell_styled(row, lc, label, lbl_font, lbl_fill, r_align)
        cell_styled(row, vc, val or '', val_font, val_fill, l_align)

    def lv2(row, l1, v1, l2, v2):
        lv(row, l1, v1, 2, 3)
        cell_styled(row, 4, l2 or '', lbl_font, lbl_fill, r_align)
        cell_styled(row, 5, v2 or '', val_font, val_fill, l_align)
        cell_styled(row, 6, '', lbl_fill, lbl_fill, c_align)

    # ── Title ──
    ws.row_dimensions[r].height = 36
    proj_code = safe_str(form.get('ProjectCode'))
    phase     = safe_str(form.get('Phase'))
    hrow(r, 1, 6,
         'HW Gaming 加速實驗室  |  測試需求資訊  —  ' + proj_code + (' ' + phase if phase else ''),
         hdr_font, hdr_fill, c_align)
    r += 1

    ws.row_dimensions[r].height = 5; r += 1  # spacer

    # ── Project Info ──
    ws.row_dimensions[r].height = 22
    hrow(r, 1, 6, ' PROJECT INFORMATION', cat_font, cat_fill,
         Alignment(horizontal='left', vertical='center'))
    r += 1

    fields = [
        ('Project Code',   safe_str(form.get('ProjectCode')),
         'Phase',          safe_str(form.get('Phase'))),
        ('Project Name',   safe_str(form.get('ProjectName')),
         'Product Line',   safe_str(form.get('ProductLine'))),
        ('Sub-series',     safe_str(form.get('SubSeries')),
         'Schedule Start', safe_str(form.get('ScheduleStart') or '')[:10]),
        ('Requester',      safe_str(form.get('Requester')),
         'Requester Email',safe_str(form.get('RequesterEmail'))),
        ('EE Contact',     safe_str(form.get('EEContact')),
         'Submit Date',    safe_str(form.get('SubmittedDate') or '')[:10]),
    ]
    for l1, v1, l2, v2 in fields:
        ws.row_dimensions[r].height = 20
        lv2(r, l1, v1, l2, v2)
        r += 1

    ws.row_dimensions[r].height = 5; r += 1  # spacer

    # ── SKU Table ──
    ws.row_dimensions[r].height = 22
    hrow(r, 1, 6, ' SKU 機台清單', cat_font, cat_fill,
         Alignment(horizontal='left', vertical='center'))
    r += 1

    sku_hdrs = ['#', 'SKU 編號', 'CPU', 'GPU', 'Panel', 'DDR / SSD']
    ws.row_dimensions[r].height = 20
    for ci, h in enumerate(sku_hdrs, 1):
        cell_styled(r, ci, h, tbl_font, cat_fill, c_align)
    r += 1

    skus = parse_sku_list(safe_str(form.get('SKUList')))
    for i, sku in enumerate(skus):
        ws.row_dimensions[r].height = 20
        bg = row_even if i % 2 == 0 else row_odd
        data = [str(i+1), sku['name'], sku['cpu'], sku['gpu'], sku['panel'], sku['ddr']]
        for ci, v in enumerate(data, 1):
            bold = (ci == 2)
            cell_styled(r, ci, v,
                        Font(name='微軟正黑體', size=10, bold=bold),
                        bg, c_align)
        r += 1

    ws.row_dimensions[r].height = 5; r += 1  # spacer

    # ── Report Scope ──
    ws.row_dimensions[r].height = 22
    hrow(r, 1, 6, ' REPORT SCOPE', cat_font, cat_fill,
         Alignment(horizontal='left', vertical='center'))
    r += 1

    ws.row_dimensions[r].height = 20
    cell_styled(r, 1, '',       lbl_font, lbl_fill, c_align)
    cell_styled(r, 2, 'Report', lbl_font, lbl_fill, c_align)
    cell_styled(r, 3, '指定 SKU', lbl_font, lbl_fill, c_align)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    cell_styled(r, 4, '測試項目', lbl_font, lbl_fill, c_align)
    r += 1

    scope_data = [
        ('ACC',            split_csv(form.get('ReportACC')),    form.get('ACCItems',     '')),
        ('TPR',            split_csv(form.get('ReportTPR')),    form.get('TPRItems',     '')),
        ('SOW Performance',split_csv(form.get('ReportSOW')),    form.get('SOWItems',     '')),
        ('HW Information', split_csv(form.get('ReportHWInfo')), form.get('HWInfoItems',  '')),
    ]
    for idx2, (name, skus_list, items_text) in enumerate(scope_data):
        if not skus_list: continue
        ws.row_dimensions[r].height = 20
        bg = row_even if idx2 % 2 == 0 else row_odd
        cell_styled(r, 1, '',                      dat_font, bg, c_align)
        cell_styled(r, 2, name,                    Font(name='微軟正黑體', size=11, bold=True), bg, c_align)
        cell_styled(r, 3, ', '.join(skus_list),    dat_font, bg, l_align)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        items_str = ' | '.join(parse_items(items_text))
        cell_styled(r, 4, items_str, dat_font, bg, l_align)
        r += 1

    # ── Notes ──
    ws.row_dimensions[r].height = 5; r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    note = ws.cell(row=r, column=1)
    note.value = '⚠ 注意：底板 sheets 內含範本數據，請在填入新測試結果前先清除既有數值。圖表因工具限制無法自動複製，如需請手動調整。'
    note.font  = Font(name='微軟正黑體', size=10, italic=True, color='FF5C4400')
    note.fill  = PatternFill('solid', fgColor='FFFCC0')
    note.alignment = l_align
    note.border = bdr
    ws.row_dimensions[r].height = 30

    return ws


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("GL HW Gaming Lab — Test Report Template Generator")
        print("=" * 60)
        print()
        print("Usage:")
        print("  python generate_template.py <form_data.json> [output.xlsx]")
        print()
        print("Example:")
        print("  python generate_template.py FX611H_PR1.json")
        print()
        sys.exit(1)

    json_path = Path(sys.argv[1])
    if not json_path.exists():
        print(f"Error: file not found — {json_path}")
        sys.exit(1)

    with open(json_path, 'r', encoding='utf-8') as f:
        form = json.load(f)

    proj_code = safe_str(form.get('ProjectCode')) or 'UNKNOWN'
    phase     = safe_str(form.get('Phase')) or ''
    ts        = datetime.now().strftime('%y%m%d')

    if len(sys.argv) > 2:
        output_path = Path(sys.argv[2])
    else:
        fname = f"{proj_code}_{phase}_RESULT_TEMPLATE_{ts}.xlsx"
        output_path = json_path.parent / fname

    print()
    print(f"  Project : {proj_code}  {phase}")
    print(f"  Output  : {output_path}")
    print()

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Cover
    build_cover_sheet(wb, form, idx=0)
    print("  ✓ 需求資訊 (Cover)")

    # 2. ACC
    acc_skus = split_csv(form.get('ReportACC'))
    if acc_skus:
        items = parse_items(form.get('ACCItems', ''))
        try:
            acc_wb = openpyxl.load_workbook(TEMPLATE_FILES['acc'])
            for item in items:
                src = ACC_SHEET_MAP.get(item)
                if src and src in acc_wb.sheetnames:
                    title = f'ACC｜{item}'
                    copy_worksheet(acc_wb[src], wb, title)
                    print(f"  ✓ {title}")
                else:
                    print(f"  ⚠  ACC: '{item}' not found in template (skip)")
        except FileNotFoundError:
            print(f"  ✗  ACC template not found: {TEMPLATE_FILES['acc']}")

    # 3. TPR
    tpr_skus = split_csv(form.get('ReportTPR'))
    if tpr_skus:
        tpr = parse_tpr(form.get('TPRItems', ''))
        try:
            tpr_wb = openpyxl.load_workbook(TEMPLATE_FILES['tpr'])
            for key, sheet_name in TPR_SHEET_MAP.items():
                if tpr[key] and sheet_name in tpr_wb.sheetnames:
                    title = f'TPR｜{sheet_name}'
                    copy_worksheet(tpr_wb[sheet_name], wb, title)
                    print(f"  ✓ {title}")
        except FileNotFoundError:
            print(f"  ✗  TPR template not found: {TEMPLATE_FILES['tpr']}")

    # 4. SOW
    sow_skus = split_csv(form.get('ReportSOW'))
    if sow_skus:
        try:
            sow_wb = openpyxl.load_workbook(TEMPLATE_FILES['sow'])
            src = SOW_SHEET_NAME if SOW_SHEET_NAME in sow_wb.sheetnames else sow_wb.active.title
            copy_worksheet(sow_wb[src], wb, 'SOW｜Performance')
            print("  ✓ SOW｜Performance")
        except FileNotFoundError:
            print(f"  ✗  SOW template not found: {TEMPLATE_FILES['sow']}")

    # 5. HW Info
    hwi_skus = split_csv(form.get('ReportHWInfo'))
    if hwi_skus:
        try:
            hwi_wb = openpyxl.load_workbook(TEMPLATE_FILES['hwinfo'])
            src = hwi_wb.sheetnames[HWINFO_SHEET_IDX]
            copy_worksheet(hwi_wb[src], wb, 'HW Info｜Summary')
            print("  ✓ HW Info｜Summary")
        except FileNotFoundError:
            print(f"  ✗  HW Info template not found: {TEMPLATE_FILES['hwinfo']}")

    wb.save(output_path)
    print()
    print(f"  ✅ 完成！已儲存至：{output_path}")
    print()


if __name__ == '__main__':
    main()

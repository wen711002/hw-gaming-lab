"""
產生 3 份模擬 EE 測試結果報告 Excel（含詳細 SKU / GPU 資訊）
供 result.html 下載測試使用
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"D:\GL HTML\HW_Gaming_Lab\mock_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

C_NAVY  = "1F3864"
C_BLUE  = "2E74B5"
C_GRAY1 = "D9E1F2"
C_GRAY2 = "F2F2F2"
C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_PURPLE = "7030A0"
C_GREEN  = "70AD47"
C_AMBER  = "FFC000"

def bd_s(style="thin", color="000000"):
    return Side(style=style, color=color)

def full_border(color="000000"):
    s = bd_s("thin", color)
    return Border(left=s, right=s, top=s, bottom=s)

def sc(cell, v=None, bold=False, size=10, fg=C_BLACK, bg=None,
       h="left", wt=False, border=False):
    if v is not None:
        cell.value = v
    cell.font = Font(name="Arial", bold=bold, size=size, color=fg)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wt)
    if border:
        cell.border = full_border()

# ── Cover ──────────────────────────────────────────────────────────────────────
def make_cover(ws, proj):
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 8

    ws.merge_cells("B1:E1")
    sc(ws["B1"], "HW Gaming Lab — 效能測試結果報告",
       bold=True, size=14, fg=C_WHITE, bg=C_NAVY, h="center")
    ws.row_dimensions[1].height = 36

    info = [
        ("Project Name",  proj["name"]),
        ("Project Code",  proj["code"]),
        ("Phase",         proj["phase"]),
        ("Platform",      proj["platform"]),
        ("OS Version",    proj["os"]),
        ("Test Engineer", proj["engineer"]),
        ("Test Date",     proj["date"]),
        ("Report Ver.",   proj["ver"]),
    ]
    for i, (label, val) in enumerate(info, start=3):
        sc(ws[f"B{i}"], label, bold=True, bg=C_GRAY1, border=True)
        sc(ws[f"C{i}"], val, border=True)
        ws.merge_cells(f"C{i}:E{i}")

    row = 3 + len(info) + 2
    ws.merge_cells(f"B{row}:E{row}")
    sc(ws[f"B{row}"], "SKU 清單", bold=True, size=11, fg=C_WHITE, bg=C_BLUE, h="center")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, label in zip(["B","C","D","E"], ["BD / 料號", "CPU", "GPU", "Panel / RAM"]):
        sc(ws[f"{col}{row}"], label, bold=True, bg=C_GRAY2, border=True, h="center")
    row += 1

    for sku in proj["skus"]:
        for col, val in zip(["B","C","D","E"], sku):
            sc(ws[f"{col}{row}"], val, border=True, h="left")
        row += 1

    ws.sheet_properties.tabColor = C_NAVY

# ── 通用 Sheet 建立器 ──────────────────────────────────────────────────────────
def make_data_sheet(ws, sku_names, items, data_map, col_widths,
                    tab_color, item_col_width=44):
    """
    sku_names : list of str — 欄標題（每個 SKU 一欄）
    items     : list of str — 測試項目名稱（依序）
    data_map  : dict[sku_name][item_index] = value
    """
    max_col = 2 + len(sku_names)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sc(ws.cell(1, 1), "▶  SKU：" + "   |   ".join(sku_names),
       bold=True, fg=C_WHITE, bg=C_BLUE)
    ws.row_dimensions[1].height = 24

    for ci, h in enumerate(["No.", "Test Item"] + sku_names, 1):
        sc(ws.cell(2, ci), h, bold=True, bg=C_GRAY1, border=True, h="center")

    for ri, item in enumerate(items, 1):
        row = ri + 2
        sc(ws.cell(row, 1), ri, h="center", border=True)
        sc(ws.cell(row, 2), item, border=True)
        for ci, sku in enumerate(sku_names):
            val = data_map.get(sku, {}).get(ri - 1)
            sc(ws.cell(row, ci + 3), val if val is not None else "N/A",
               h="center", border=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = item_col_width
    for ci in range(len(sku_names)):
        ws.column_dimensions[get_column_letter(ci + 3)].width = col_widths

    ws.sheet_properties.tabColor = tab_color

# ── HW Info ────────────────────────────────────────────────────────────────────
def make_hwinfo(ws, proj):
    sku = proj["skus"][0]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 38

    ws.merge_cells("A1:B1")
    sc(ws["A1"], "HW Info — 系統規格摘要",
       bold=True, size=12, fg=C_WHITE, bg=C_PURPLE, h="center")
    ws.row_dimensions[1].height = 28

    rows = [
        ("Project Code",      proj["code"]),
        ("Phase",             proj["phase"]),
        ("BD / 料號",          sku[0]),
        ("CPU",               sku[1]),
        ("GPU",               sku[2]),
        ("Panel",             proj.get("panel", "15.6\" FHD 144Hz")),
        ("RAM",               proj.get("ram",   "16GB DDR5-5600")),
        ("Storage",           proj.get("storage","1TB PCIe 4.0 NVMe")),
        ("Battery",           proj.get("batt_cap","90Wh")),
        ("OS",                proj["os"]),
        ("BIOS Ver.",         proj.get("bios","301")),
        ("EC Ver.",           proj.get("ec",  "101")),
        ("GPU Driver",        proj.get("gpu_drv","572.83")),
        ("LAN Driver",        proj.get("lan_drv","12.19.2.60")),
        ("Audio Driver",      proj.get("audio_drv","6.0.9550.1")),
        ("Turbo Mode TDP (W)",proj.get("tdp_turbo","45")),
        ("Standard TDP (W)",  proj.get("tdp_std","35")),
        ("Fan Mode tested",   proj.get("fan_mode","Turbo")),
    ]
    for ri, (label, val) in enumerate(rows, 3):
        sc(ws[f"A{ri}"], label, bold=True, bg=C_GRAY2, border=True)
        sc(ws[f"B{ri}"], str(val), border=True)

    ws.sheet_properties.tabColor = C_PURPLE

# ── 主建立函式 ─────────────────────────────────────────────────────────────────
def create_report(proj, filename):
    wb = Workbook()
    wb.active.title = "_tmp"

    ws = wb.create_sheet("Cover")
    make_cover(ws, proj)

    reports = proj["reports"]

    # ── ACC ──
    if "acc" in reports:
        acc_skus = [s[0] for s in proj["skus"]]

        ws = wb.create_sheet("ACC｜Cinebench R23 30-loop")
        make_data_sheet(
            ws, acc_skus,
            items=[
                "Cinebench R23 30-loop — Multi-Core Median (pts)",
                "Cinebench R23 30-loop — Multi-Core Max (pts)",
                "Cinebench R23 30-loop — Multi-Core Min (pts)",
                "Cinebench R23 30-loop — Single-Core Median (pts)",
                "Cinebench R23 30-loop — Single-Core Max (pts)",
                "Cinebench R23 30-loop — Single-Core Min (pts)",
                "Power Limit — Turbo Mode (W)",
                "Power Limit — Standard Mode (W)",
            ],
            data_map=proj.get("acc_cb", {}),
            col_widths=16, tab_color="4472C4", item_col_width=52
        )

        ws = wb.create_sheet("ACC｜3DMark Benchmark")
        make_data_sheet(
            ws, acc_skus,
            items=[
                "Fire Strike — Graphics Score",
                "Fire Strike — Physics Score",
                "Fire Strike Extreme — Graphics Score",
                "Fire Strike Ultra — Graphics Score",
                "Time Spy — Graphics Score",
                "Time Spy — CPU Score",
                "Time Spy Extreme — Graphics Score",
                "Port Royal — Score",
                "Speed Way — Score",
                "DLSS Feature Test (DLSS 4 MFG)",
            ],
            data_map=proj.get("acc_3dm", {}),
            col_widths=16, tab_color="4472C4", item_col_width=46
        )

    # ── TPR ──
    if "tpr" in reports:
        tpr_skus = [s[0] for s in proj["skus"]]

        ws = wb.create_sheet("TPR｜Battery life")
        make_data_sheet(
            ws, tpr_skus,
            items=[
                "MobileMark 25 — Modern Office (min)",
                "MobileMark 25 — Video Playback (min)",
                "MobileMark 25 — Content Creation (min)",
                "PCMark 10 Battery — Modern Office (min)",
                "PCMark 10 Battery — Video Playback (min)",
                "Web Browsing (Edge / YouTube) (min)",
                "Battery Capacity (Wh)",
                "Charging Time: 0→100% (min)",
            ],
            data_map=proj.get("tpr_bat", {}),
            col_widths=16, tab_color="FFC000", item_col_width=46
        )

        ws = wb.create_sheet("TPR｜Performance")
        make_data_sheet(
            ws, tpr_skus,
            items=[
                "Cinebench R23 — Multi-Core (pts)",
                "Cinebench R23 — Single-Core (pts)",
                "Cinebench R24 — Multi-Core (pts)",
                "Cinebench R24 — Single-Core (pts)",
                "Cinebench 2024 — Multi-Core (pts)",
                "PCMark 10 — Overall (score)",
                "PCMark 10 — Productivity (score)",
                "PCMark 10 — Digital Content Creation (score)",
                "CrossMark — Overall (score)",
                "CrossMark — Productivity (score)",
                "3DMark CPU Profile — 1T (pts)",
                "3DMark CPU Profile — Max Threads (pts)",
            ],
            data_map=proj.get("tpr_perf", {}),
            col_widths=16, tab_color="FFC000", item_col_width=50
        )

    # ── SOW ──
    if "sow" in reports:
        # SOW 按 GPU 分組，每個 GPU SKU 一欄
        sow_skus = [s[0] for s in proj["sow_skus"]]

        ws = wb.create_sheet("SOW｜Performance test")

        # 先寫 GPU 資訊說明行（row 1 SKU header 後多加一列 GPU 標示）
        max_col = 2 + len(sow_skus)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        sc(ws.cell(1, 1), "▶  SKU：" + "   |   ".join(sow_skus),
           bold=True, fg=C_WHITE, bg=C_BLUE)
        ws.row_dimensions[1].height = 24

        # Row 2: column headers with GPU label
        ws.cell(2, 1).value = "No."
        ws.cell(2, 2).value = "Test Item"
        for ci, sku_info in enumerate(proj["sow_skus"], 1):
            cell = ws.cell(2, ci + 2)
            cell.value = sku_info[0]  # SKU 料號
            sc(cell, bold=True, bg=C_GRAY1, border=True, h="center")
        sc(ws.cell(2, 1), bold=True, bg=C_GRAY1, border=True, h="center")
        sc(ws.cell(2, 2), bold=True, bg=C_GRAY1, border=True, h="center")

        # Row 3: GPU sub-header
        ws.cell(3, 1).value = ""
        ws.cell(3, 2).value = "GPU →"
        sc(ws.cell(3, 2), bold=True, bg="F2F2F2", border=True, h="right",
           fg="595959", size=9)
        for ci, sku_info in enumerate(proj["sow_skus"], 1):
            cell = ws.cell(3, ci + 2)
            cell.value = sku_info[2]  # GPU 名稱
            sc(cell, bold=True, bg="FFF2CC", border=True, h="center",
               fg="7F6000", size=9)
        ws.row_dimensions[3].height = 18

        # 3DMark section
        sow_3dm_items = [
            "3DMark Fire Strike — Graphics Score",
            "3DMark Fire Strike — Physics Score",
            "3DMark Fire Strike Extreme — Graphics Score",
            "3DMark Fire Strike Ultra — Graphics Score",
            "3DMark Time Spy — Graphics Score",
            "3DMark Time Spy — CPU Score",
            "3DMark Time Spy Extreme — Graphics Score",
            "3DMark Port Royal — Score",
            "3DMark Speed Way — Score",
        ]
        # Gaming FPS section
        sow_fps_items = [
            "Cyberpunk 2077 — 1080p Ultra RT (fps)",
            "Cyberpunk 2077 — 1080p Ultra RT + DLSS Quality (fps)",
            "Cyberpunk 2077 — 1440p Ultra RT (fps)",
            "Cyberpunk 2077 — 1440p Ultra RT + DLSS Quality (fps)",
            "Alan Wake 2 — 1080p Ultra RT (fps)",
            "Alan Wake 2 — 1080p Ultra RT + DLSS Quality (fps)",
            "F1 2024 — 1080p Ultra (fps)",
            "F1 2024 — 1440p Ultra (fps)",
            "CS2 — 1080p High (fps)",
            "DOTA 2 — 1080p Ultra (fps)",
            "Red Dead Redemption 2 — 1080p Ultra (fps)",
            "Elden Ring — 1080p Max Settings (fps)",
            "Black Myth: Wukong — 1080p Epic (fps)",
            "Black Myth: Wukong — 1080p Epic + DLSS Quality (fps)",
        ]

        data_3dm = proj.get("sow_3dm", {})
        data_fps = proj.get("sow_fps", {})

        row = 4
        # 3DMark header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        sc(ws.cell(row, 1), "▌ 3DMark Benchmark",
           bold=True, fg=C_WHITE, bg="2E74B5", size=11)
        ws.row_dimensions[row].height = 20
        row += 1

        for ri, item in enumerate(sow_3dm_items, 1):
            sc(ws.cell(row, 1), ri, h="center", border=True)
            sc(ws.cell(row, 2), item, border=True)
            for ci, sku_info in enumerate(proj["sow_skus"], 1):
                val = data_3dm.get(sku_info[0], {}).get(ri - 1)
                sc(ws.cell(row, ci + 2), val if val is not None else "N/A",
                   h="center", border=True)
            row += 1

        # Gaming FPS header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        sc(ws.cell(row, 1), "▌ Gaming FPS",
           bold=True, fg=C_WHITE, bg="70AD47", size=11)
        ws.row_dimensions[row].height = 20
        row += 1

        for ri, item in enumerate(sow_fps_items, 1):
            sc(ws.cell(row, 1), ri, h="center", border=True)
            sc(ws.cell(row, 2), item, border=True)
            for ci, sku_info in enumerate(proj["sow_skus"], 1):
                val = data_fps.get(sku_info[0], {}).get(ri - 1)
                sc(ws.cell(row, ci + 2), val if val is not None else "N/A",
                   h="center", border=True)
            row += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 52
        for ci in range(len(sow_skus)):
            ws.column_dimensions[get_column_letter(ci + 3)].width = 20
        ws.sheet_properties.tabColor = "70AD47"

    # ── HW Info ──
    if "hwi" in reports:
        ws = wb.create_sheet("HW Info｜Summary")
        make_hwinfo(ws, proj)

    wb.remove(wb["_tmp"])
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  OK {filename}  ({os.path.getsize(path)//1024} KB)")


# ══════════════════════════════════════════════════════════════════════════════
#  3 份模擬報告資料
# ══════════════════════════════════════════════════════════════════════════════

# ──────────────────────────────────────────────────────────────────────────────
# 1. FX611H  PR1  Intel Arrow Lake-H  3 SKU(不同GPU)  ACC + TPR + SOW
# ──────────────────────────────────────────────────────────────────────────────
#
# SKU 對應關係：
#   AQ001  Core Ultra 7 265H  +  RTX 5060 8GB   (Turbo TDP 45W)
#   AQ002  Core Ultra 7 265H  +  RTX 5070 Ti 12GB  (Turbo TDP 55W)
#   AQ003  Core Ultra 5 235H  +  RTX 5060 8GB   (Turbo TDP 45W)
#
proj1 = dict(
    name="FX611H", code="FX611H", phase="PR1",
    platform="Intel Core Ultra 200H (Arrow Lake-H) + NVIDIA GeForce RTX 50 Series",
    os="Windows 11 24H2 (Build 26100)",
    engineer="Tom Chen", date="2026-04-10", ver="v1.0",
    panel='15.6" FHD 144Hz IPS', ram="16GB DDR5-5600 (SO-DIMM×2)",
    storage="1TB WD SN850X PCIe 4.0 NVMe",
    batt_cap="90Wh", bios="FX611H.310", ec="FX611H.110",
    gpu_drv="572.83 (WHQL)", lan_drv="12.19.2.60", audio_drv="6.0.9550.1",
    tdp_turbo="45 / 55 (SKU AQ002)", tdp_std="35",
    fan_mode="Turbo Mode (Manual)",
    reports=["acc", "tpr", "sow"],
    # Cover 用的 SKU 表（BD/料號, CPU, GPU, Panel/RAM）
    skus=[
        ["FX611H-AQ001", "Intel Core Ultra 7 265H",  "RTX 5060 8GB",    '15.6" FHD 144Hz / 16GB DDR5-5600'],
        ["FX611H-AQ002", "Intel Core Ultra 7 265H",  "RTX 5070 Ti 12GB",'15.6" FHD 144Hz / 16GB DDR5-5600'],
        ["FX611H-AQ003", "Intel Core Ultra 5 235H",  "RTX 5060 8GB",    '15.6" FHD 144Hz / 16GB DDR5-5600'],
    ],
    # ACC — 以 SKU 料號為 key
    acc_cb={
        "FX611H-AQ001": {0:28480, 1:29120, 2:27940, 3:1945, 4:1978, 5:1912, 6:45, 7:35},
        "FX611H-AQ002": {0:28350, 1:28990, 2:27810, 3:1941, 4:1975, 5:1908, 6:55, 7:35},
        "FX611H-AQ003": {0:22140, 1:22780, 2:21560, 3:1812, 4:1841, 5:1783, 6:45, 7:35},
    },
    acc_3dm={
        "FX611H-AQ001": {0:19240, 1:32580, 2:9980,  3:5530,  4:12560, 5:14120, 6:6230,  7:8120,  8:3350,  9:"Supported"},
        "FX611H-AQ002": {0:31850, 1:32410, 2:16540, 3:9210,  4:20980, 5:14050, 6:10420, 7:13480, 8:5580,  9:"Supported"},
        "FX611H-AQ003": {0:18970, 1:22140, 2:9820,  3:5430,  4:12340, 5:9810,  6:6110,  7:7980,  8:3290,  9:"Supported"},
    },
    tpr_bat={
        "FX611H-AQ001": {0:682, 1:824, 2:598, 3:610, 4:910, 5:542, 6:90, 7:78},
        "FX611H-AQ002": {0:658, 1:801, 2:575, 3:588, 4:892, 5:521, 6:90, 7:82},
        "FX611H-AQ003": {0:715, 1:856, 2:628, 3:642, 4:942, 5:568, 6:90, 7:76},
    },
    tpr_perf={
        "FX611H-AQ001": {0:28350, 1:1932, 2:18850, 3:1278, 4:1520, 5:8124, 6:9546, 7:10280, 8:2042, 9:2156, 10:1245, 11:28460},
        "FX611H-AQ002": {0:28210, 1:1928, 2:18760, 3:1272, 4:1515, 5:8098, 6:9512, 7:10250, 8:2035, 9:2148, 10:1241, 11:28380},
        "FX611H-AQ003": {0:22120, 1:1810, 2:14890, 3:1198, 4:1418, 5:7650, 6:8920, 7: 9640, 8:1924, 9:2034, 10:1156, 11:22050},
    },
    # SOW — 每個 GPU SKU 獨立欄
    sow_skus=[
        # (料號,         CPU,                        GPU,              Panel)
        ("FX611H-AQ001", "Core Ultra 7 265H", "RTX 5060 8GB GDDR7",    '15.6" FHD 144Hz'),
        ("FX611H-AQ002", "Core Ultra 7 265H", "RTX 5070 Ti 12GB GDDR7",'15.6" FHD 144Hz'),
        ("FX611H-AQ003", "Core Ultra 5 235H", "RTX 5060 8GB GDDR7",    '15.6" FHD 144Hz'),
    ],
    sow_3dm={
        "FX611H-AQ001": {0:19240, 1:32580, 2:9980,  3:5530,  4:12560, 5:14120, 6:6230,  7:8120,  8:3350},
        "FX611H-AQ002": {0:31850, 1:32410, 2:16540, 3:9210,  4:20980, 5:14050, 6:10420, 7:13480, 8:5580},
        "FX611H-AQ003": {0:18970, 1:22140, 2:9820,  3:5430,  4:12340, 5:9810,  6:6110,  7:7980,  8:3290},
    },
    sow_fps={
        # CP 1080p Ultra RT, CP 1080p+DLSS, CP 1440p RT, CP 1440p+DLSS,
        # AW2 1080p RT, AW2 1080p+DLSS, F1 1080p, F1 1440p,
        # CS2 1080p, DOTA2 1080p, RDR2 1080p, Elden 1080p,
        # Wukong 1080p, Wukong 1080p+DLSS
        "FX611H-AQ001": {0:58,  1:98,  2:38,  3:67,  4:52,  5:88,  6:148, 7:101, 8:320, 9:210, 10:78, 11:112, 12:44, 13:76},
        "FX611H-AQ002": {0:89,  1:148, 2:62,  3:104, 4:81,  5:135, 6:198, 7:142, 8:340, 9:225, 10:105,11:148, 12:68, 13:115},
        "FX611H-AQ003": {0:55,  1:94,  2:36,  3:64,  4:50,  5:84,  6:145, 7:98,  8:312, 9:205, 10:75, 11:108, 12:42, 13:72},
    },
)

# ──────────────────────────────────────────────────────────────────────────────
# 2. FA401EA  PR2  AMD Strix Point (Ryzen AI 300)  3 SKU(不同APU/iGPU)  ACC + TPR
# ──────────────────────────────────────────────────────────────────────────────
#
# SKU 對應關係：
#   BQ001  Ryzen AI 9 HX 375  +  Radeon RX 890M (iGPU 16CU @ 2900MHz)
#   BQ002  Ryzen AI 7 HX 360  +  Radeon RX 880M (iGPU 12CU @ 2700MHz)
#   BQ003  Ryzen AI 5 360     +  Radeon RX 860M (iGPU 8CU  @ 2600MHz)
#
proj2 = dict(
    name="FA401EA", code="FA401EA", phase="PR2",
    platform="AMD Ryzen AI 300 Series (Strix Point) — Zen5 + RDNA 3.5 iGPU",
    os="Windows 11 24H2 (Build 26100)",
    engineer="Linda Wu", date="2026-04-18", ver="v1.2",
    panel='14" 2.8K 90Hz OLED', ram="32GB LPDDR5X-7500 (On-board)",
    storage="1TB Micron 4500 PCIe 4.0 NVMe",
    batt_cap="73Wh", bios="FA401EA.215", ec="FA401EA.108",
    gpu_drv="Adrenalin 24.12.1 (iGPU)", lan_drv="N/A (Wi-Fi only)",
    audio_drv="6.0.9550.1",
    tdp_turbo="28", tdp_std="15",
    fan_mode="Performance Mode",
    reports=["acc", "tpr"],
    skus=[
        ["FA401EA-BQ001", "Ryzen AI 9 HX 375",  "Radeon RX 890M (iGPU, 16CU)", '14" 2.8K 90Hz OLED / 32GB LPDDR5X-7500'],
        ["FA401EA-BQ002", "Ryzen AI 7 HX 360",  "Radeon RX 880M (iGPU, 12CU)", '14" 2.8K 90Hz OLED / 32GB LPDDR5X-7500'],
        ["FA401EA-BQ003", "Ryzen AI 5 360",      "Radeon RX 860M (iGPU, 8CU)",  '14" FHD 60Hz IPS   / 16GB LPDDR5X-6400'],
    ],
    acc_cb={
        "FA401EA-BQ001": {0:24580, 1:25120, 2:24010, 3:1756, 4:1789, 5:1721, 6:28, 7:15},
        "FA401EA-BQ002": {0:21340, 1:21890, 2:20820, 3:1698, 4:1728, 5:1668, 6:28, 7:15},
        "FA401EA-BQ003": {0:15670, 1:16120, 2:15240, 3:1542, 4:1572, 5:1510, 6:20, 7:12},
    },
    acc_3dm={
        "FA401EA-BQ001": {0:8920,  1:14250, 2:4520, 3:2450, 4:5840, 5:5120, 6:2890, 7:"N/A", 8:"N/A", 9:"N/A"},
        "FA401EA-BQ002": {0:7650,  1:14180, 2:3890, 3:2110, 4:5010, 5:5090, 6:2480, 7:"N/A", 8:"N/A", 9:"N/A"},
        "FA401EA-BQ003": {0:5230,  1:9540,  2:2640, 3:1430, 4:3420, 5:3680, 6:1690, 7:"N/A", 8:"N/A", 9:"N/A"},
    },
    tpr_bat={
        "FA401EA-BQ001": {0:534, 1:698, 2:468, 3:476, 4:812, 5:612, 6:73, 7:56},
        "FA401EA-BQ002": {0:568, 1:734, 2:498, 3:508, 4:848, 5:648, 6:73, 7:54},
        "FA401EA-BQ003": {0:612, 1:789, 2:542, 3:548, 4:901, 5:698, 6:73, 7:52},
    },
    tpr_perf={
        "FA401EA-BQ001": {0:24560, 1:1748, 2:16320, 3:1189, 4:1398, 5:7456, 6:8890, 7:9540, 8:1856, 9:1945, 10:980, 11:24480},
        "FA401EA-BQ002": {0:21230, 1:1690, 2:14150, 3:1145, 4:1350, 5:7123, 6:8542, 7:9180, 8:1798, 9:1882, 10:948, 11:21160},
        "FA401EA-BQ003": {0:15640, 1:1534, 2:10420, 3:1042, 4:1228, 5:6780, 6:8102, 7:8620, 8:1645, 9:1724, 10:872, 11:15580},
    },
)

# ──────────────────────────────────────────────────────────────────────────────
# 3. G513RM  PR1  AMD Dragon Range  2 SKU(不同GPU)  TPR + SOW + HW Info
# ──────────────────────────────────────────────────────────────────────────────
#
# SKU 對應關係：
#   HQ001  Ryzen 9 7945HX  +  RTX 4070 8GB  (Turbo TDP 50W)
#   HQ002  Ryzen 9 7945HX  +  RTX 4060 8GB  (Turbo TDP 45W)
#
proj3 = dict(
    name="G513RM", code="G513RM", phase="PR1",
    platform="AMD Ryzen 9 7945HX (Dragon Range, Zen4) + NVIDIA GeForce RTX 40 Series",
    os="Windows 11 23H2 (Build 22631)",
    engineer="Kevin Lee", date="2026-04-25", ver="v1.0",
    panel='15.6" QHD 165Hz IPS', ram="32GB DDR5-4800 (SO-DIMM×2)",
    storage="1TB Samsung 990 Pro PCIe 4.0 NVMe",
    batt_cap="90Wh", bios="G513RM.318", ec="G513RM.115",
    gpu_drv="572.60 (WHQL)", lan_drv="12.19.2.60", audio_drv="6.0.9550.1",
    tdp_turbo="50 / 45 (SKU HQ002)", tdp_std="35",
    fan_mode="Turbo Mode (Manual)",
    reports=["tpr", "sow", "hwi"],
    skus=[
        ["G513RM-HQ001", "Ryzen 9 7945HX", "RTX 4070 8GB GDDR6", '15.6" QHD 165Hz / 32GB DDR5-4800'],
        ["G513RM-HQ002", "Ryzen 9 7945HX", "RTX 4060 8GB GDDR6", '15.6" QHD 165Hz / 32GB DDR5-4800'],
    ],
    tpr_bat={
        "G513RM-HQ001": {0:420, 1:562, 2:385, 3:378, 4:645, 5:412, 6:90, 7:85},
        "G513RM-HQ002": {0:445, 1:585, 2:408, 3:402, 4:668, 5:435, 6:90, 7:82},
    },
    tpr_perf={
        "G513RM-HQ001": {0:22450, 1:1734, 2:14980, 3:1178, 4:1388, 5:7234, 6:8678, 7:9340, 8:1812, 9:1908, 10:1045, 11:22380},
        "G513RM-HQ002": {0:22380, 1:1730, 2:14920, 3:1174, 4:1382, 5:7218, 6:8654, 7:9312, 8:1808, 9:1902, 10:1040, 11:22310},
    },
    sow_skus=[
        ("G513RM-HQ001", "Ryzen 9 7945HX", "RTX 4070 8GB GDDR6", '15.6" QHD 165Hz'),
        ("G513RM-HQ002", "Ryzen 9 7945HX", "RTX 4060 8GB GDDR6", '15.6" QHD 165Hz'),
    ],
    sow_3dm={
        "G513RM-HQ001": {0:20420, 1:21580, 2:10580, 3:5890, 4:13240, 5:10820, 6:6590, 7:9480, 8:3820},
        "G513RM-HQ002": {0:14560, 1:21420, 2: 7420, 3:4120, 4: 9340, 5:10780, 6:4680, 7:6620, 8:2680},
    },
    sow_fps={
        # CP 1080p RT, CP 1080p+DLSS, CP 1440p RT, CP 1440p+DLSS,
        # AW2 1080p RT, AW2 1080p+DLSS, F1 1080p, F1 1440p,
        # CS2 1080p, DOTA2 1080p, RDR2 1080p, Elden 1080p,
        # Wukong 1080p, Wukong 1080p+DLSS
        "G513RM-HQ001": {0:52,  1:88,  2:34,  3:58,  4:48,  5:80,  6:138, 7:94,  8:298, 9:198, 10:92, 11:108, 12:38, 13:65},
        "G513RM-HQ002": {0:38,  1:66,  2:25,  3:44,  4:35,  5:62,  6:118, 7:80,  8:268, 9:178, 10:75, 11: 90, 12:28, 13:50},
    },
)


# ══════════════════════════════════════════════════════════════════════════════
print("\nGL HW Gaming Lab — 產生模擬測試結果報告")
print("=" * 52)
create_report(proj1, "FX611H_PR1_Result_20260410.xlsx")
create_report(proj2, "FA401EA_PR2_Result_20260418.xlsx")
create_report(proj3, "G513RM_PR1_Result_20260425.xlsx")
print(f"\n完成！輸出目錄：{OUTPUT_DIR}")

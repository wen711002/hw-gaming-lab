import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))

NAVY = "1F2D4E"
DARK = "2E4057"
GRAY = "D9D9D9"
WHITE = "FFFFFF"

def hdr(ws, row, col, value, bg=NAVY):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def plain(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold)
    c.alignment = Alignment(vertical="center")
    return c

def make_file(fname, project_code, phase, os_ver,
              bd1, bd2, cpu1, cpu2, panel1, panel2, ddr1, ddr2,
              bat, bench, fps):

    wb = Workbook()

    # ── 說明 ──────────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "說明"
    c = ws0.cell(row=1, column=1, value="HW Gaming Lab 效能測試結果報告")
    c.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.merge_cells("A1:E1")
    ws0.row_dimensions[1].height = 30
    ws0.cell(row=2, column=1, value="此為系統模擬測試資料").font = Font(name="Arial", italic=True)

    # ── Performance Test ──────────────────────────────────────────────────
    ws = wb.create_sheet("Performance Test")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Row 1 title
    ws.merge_cells("A1:E1")
    hdr(ws, 1, 1, "HW Gaming Lab 效能測試結果")
    ws.row_dimensions[1].height = 28

    # Row 2-3 project info
    plain(ws, 2, 1, "Project Code:", bold=True)
    plain(ws, 2, 4, project_code)
    plain(ws, 3, 1, "OS / Phase:", bold=True)
    plain(ws, 3, 4, os_ver)
    plain(ws, 3, 5, phase)

    # Row 5 column headers
    for col, val in enumerate(["No.", "Test Item", "Unit", "SKU-01", "SKU-02"], 1):
        c = ws.cell(row=5, column=col, value=val)
        c.font = Font(name="Arial", bold=True)
        c.fill = PatternFill("solid", start_color=GRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6-9 SKU specs
    specs = [("BD", bd1, bd2), ("CPU", cpu1, cpu2), ("Panel", panel1, panel2), ("DDR", ddr1, ddr2)]
    for i, (label, v1, v2) in enumerate(specs, 6):
        plain(ws, i, 2, label, bold=True)
        plain(ws, i, 4, v1)
        plain(ws, i, 5, v2)

    # Battery section
    ws.merge_cells("A11:E11")
    hdr(ws, 11, 1, "🔋 Battery Life", bg=DARK)
    ws.row_dimensions[11].height = 24
    bat_items = [
        "MobileMark 25 (Modern Office)",
        "MobileMark 25 (Video Playback)",
        "PCMark 10 Battery (Modern Office)",
        "PCMark 10 Battery (Video)",
    ]
    for i, (item, v1, v2) in enumerate(zip(bat_items, bat[::2], bat[1::2]), 12):
        plain(ws, i, 1, i - 11)
        plain(ws, i, 2, item)
        plain(ws, i, 3, "hr")
        plain(ws, i, 4, v1)
        plain(ws, i, 5, v2)

    # Benchmark section
    ws.merge_cells("A20:E20")
    hdr(ws, 20, 1, "📊 Benchmark", bg=DARK)
    ws.row_dimensions[20].height = 24
    bench_items = [
        ("Cinebench R24 (Multi-Core)", "pts"),
        ("Cinebench R24 (Single-Core)", "pts"),
        ("3DMark TimeSpy", "score"),
        ("3DMark FireStrike", "score"),
        ("PCMark 10", "score"),
    ]
    for i, ((item, unit), v1, v2) in enumerate(zip(bench_items, bench[::2], bench[1::2]), 21):
        plain(ws, i, 1, i - 20)
        plain(ws, i, 2, item)
        plain(ws, i, 3, unit)
        plain(ws, i, 4, v1)
        plain(ws, i, 5, v2)

    # Gaming FPS section
    ws.merge_cells("A40:E40")
    hdr(ws, 40, 1, "🎮 Gaming FPS", bg=DARK)
    ws.row_dimensions[40].height = 24
    fps_items = [
        "Cyberpunk 2077 (1080p Ultra)",
        "Cyberpunk 2077 (1440p Ultra)",
        "Elden Ring (1080p Max)",
        "CS2 (1080p High)",
        "DOTA2 (1080p Ultra)",
    ]
    for i, (item, v1, v2) in enumerate(zip(fps_items, fps[::2], fps[1::2]), 41):
        plain(ws, i, 1, i - 40)
        plain(ws, i, 2, item)
        plain(ws, i, 3, "fps")
        plain(ws, i, 4, v1)
        plain(ws, i, 5, v2)

    ws.cell(row=71, column=1, value="備註：此為模擬資料，僅供系統測試用途。").font = Font(name="Arial", italic=True, color="808080")

    # ── SKU Table ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("SKU Table")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18

    ws2.merge_cells("A1:E1")
    hdr(ws2, 1, 1, "SKU 規格清單")
    ws2.row_dimensions[1].height = 28

    plain(ws2, 4, 1, "Project:", bold=True); plain(ws2, 4, 2, project_code)
    plain(ws2, 5, 1, "Phase:", bold=True);   plain(ws2, 5, 2, phase)
    plain(ws2, 6, 1, "Test Date:", bold=True); plain(ws2, 6, 2, "2026-04-23")

    for col, val in enumerate(["BD", "SKU", "GPU", "Panel", "DDR"], 1):
        c = ws2.cell(row=11, column=col, value=val)
        c.font = Font(name="Arial", bold=True)
        c.fill = PatternFill("solid", start_color=GRAY)
        c.alignment = Alignment(horizontal="center")

    sku_data = [
        (bd1, bd1, cpu1.replace("Intel Core Ultra ","iCU").replace("AMD Ryzen ","R"), panel1, ddr1),
        (bd2, bd2, cpu2.replace("Intel Core Ultra ","iCU").replace("AMD Ryzen ","R"), panel2, ddr2),
    ]
    for r, row in enumerate(sku_data, 12):
        for c, val in enumerate(row, 1):
            plain(ws2, r, c, val)

    path = os.path.join(OUT, fname)
    wb.save(path)
    print(f"Saved: {path}")


# ── File 1: G814JV EVT ────────────────────────────────────────────────────
make_file(
    "G814JV_EVT_RESULT_NEW_260423.xlsx",
    "G814JV", "EVT", "Win11 24H2",
    "G814JV-0021", "G814JV-0032",
    "Intel Core Ultra 9 185HX", "Intel Core Ultra 7 165HX",
    '18" QHD 240Hz', '18" FHD 144Hz',
    "64GB DDR5-5600", "32GB DDR5-5200",
    bat=[8.5, 7.9, 11.2, 10.8, 9.1, 8.6, 12.3, 11.9],
    bench=[18542, 17890, 2156, 2089, 12450, 11980, 28930, 27650, 7823, 7541],
    fps=[87, 82, 54, 51, 112, 108, 245, 231, 189, 178],
)

# ── File 2: F607BA PR ─────────────────────────────────────────────────────
make_file(
    "F607BA_PR_RESULT_NEW_260423.xlsx",
    "F607BA", "PR", "Win11 23H2",
    "F607BA-0011", "F607BA-0022",
    "Intel Core Ultra 5 125H", "Intel Core Ultra 5 125H",
    '16" FHD 144Hz', '16" FHD 60Hz',
    "16GB DDR5-4800", "16GB DDR5-4800",
    bat=[10.2, 11.5, 13.8, 15.1, 11.0, 12.3, 14.5, 16.2],
    bench=[14230, 14156, 1876, 1871, 7820, 5430, 18920, 12340, 6890, 6750],
    fps=[62, 41, 38, 25, 89, 65, 178, 124, 145, 98],
)

# ── File 3: G634JY ER2 ────────────────────────────────────────────────────
make_file(
    "G634JY_ER2_RESULT_R1_260423.xlsx",
    "G634JY", "ER2", "Win11 24H2",
    "G634JY-0041", "G634JY-0052",
    "AMD Ryzen 9 7945HX", "AMD Ryzen 7 7745HX",
    '16" QHD 240Hz', '16" FHD 165Hz',
    "32GB DDR5-4800", "16GB DDR5-4800",
    bat=[7.8, 8.9, 10.5, 11.8, 8.3, 9.6, 11.2, 12.9],
    bench=[21890, 17320, 1923, 1756, 10230, 7890, 24560, 18970, 7230, 6980],
    fps=[74, 58, 46, 36, 98, 81, 215, 168, 167, 138],
)

print("All 3 mock files created.")

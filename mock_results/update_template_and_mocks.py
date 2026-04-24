import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from copy import copy

SRC   = r'G:\HW Gaming\Result\HWGamingLab_Result_Template_v5.xlsx'
LOCK  = r'G:\HW Gaming\Result\HWGamingLab_Result_Template_v5_locked.xlsx'
OUT   = os.path.dirname(os.path.abspath(__file__))

# Match actual template label style (sampled from rows 4-7)
LABEL_FONT  = Font(name='Arial', bold=True, size=9, color='FF1F3864')
LABEL_FILL  = PatternFill('solid', fgColor='FFBDD7EE')
LABEL_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
VALUE_FONT  = Font(name='Arial', size=9)
VALUE_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_label(cell, text):
    cell.value = text
    cell.font  = copy(LABEL_FONT)
    cell.fill  = copy(LABEL_FILL)
    cell.alignment = copy(LABEL_ALIGN)

def set_value(cell, text=''):
    cell.value = text
    cell.font  = copy(VALUE_FONT)
    cell.alignment = copy(VALUE_ALIGN)

def ensure_unmerged(ws, rng):
    try:
        ws.unmerge_cells(rng)
    except (KeyError, ValueError):
        pass

def ensure_merged(ws, rng):
    if not any(str(m) == rng for m in ws.merged_cells.ranges):
        ws.merge_cells(rng)

# ── 1. Update SKU Table metadata rows 9-10 ───────────────────────────────────
wb = openpyxl.load_workbook(SRC)
ws2 = wb['SKU Table']

# Clear any stale H/I data from rows 8-9 (old layout)
for r in (8, 9):
    for c in range(7, 12):
        ws2.cell(r, c).value = None

# Fix openpyxl bug: insert_rows doesn't shift merged ranges,
# so B10:F10 (BD與SKU清單 header) may still cover row 10 incorrectly.
ensure_unmerged(ws2, 'B10:F10')

# ── Row 9: 申請人 Email (B9/C9:D9) + Product Line (E9/F9) ────────────────
ensure_unmerged(ws2, 'C9:D9')
set_label(ws2.cell(9, 2), '申請人 Email')
set_value(ws2.cell(9, 3))
set_label(ws2.cell(9, 5), 'Product Line')
set_value(ws2.cell(9, 6))
ensure_merged(ws2, 'C9:D9')

# ── Row 10: Sub-series (B10/C10:D10) + 專案名稱(系統) (E10/F10) ─────────
# Insert only if not yet present
if ws2.cell(10, 2).value != 'Sub-series':
    ws2.insert_rows(10)

ensure_unmerged(ws2, 'C10:D10')
set_label(ws2.cell(10, 2), 'Sub-series')
set_value(ws2.cell(10, 3))
set_label(ws2.cell(10, 5), '專案名稱(系統)')
set_value(ws2.cell(10, 6))
ensure_merged(ws2, 'C10:D10')

# Re-merge BD與SKU清單 header at row 11 (shifted from row 10)
ensure_merged(ws2, 'B11:F11')

wb.save(SRC)
print("Template rows 9-10 updated.")

# ── 2. Re-apply protection ───────────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC)
unlocked = Protection(locked=False)
locked   = Protection(locked=True)

ws = wb['Performance Test']
CATEGORY_ROWS = {11, 20, 40}
for row in ws.iter_rows():
    r = row[0].row
    for cell in row:
        c = cell.column
        if r == 1:                       cell.protection = locked
        elif r in (2, 3):                cell.protection = unlocked
        elif r == 4:                     cell.protection = locked
        elif r == 5:
            cell.protection = locked if c <= 3 else unlocked
        elif 6 <= r <= 9:
            cell.protection = locked if c <= 3 else unlocked
        elif r == 10:                    cell.protection = locked
        elif r in CATEGORY_ROWS:         cell.protection = locked
        elif r >= 71:                    cell.protection = locked
        else:                            cell.protection = unlocked
ws.protection.sheet = True
ws.protection.password = ''
ws.protection.selectLockedCells = False
ws.protection.selectUnlockedCells = False

ws2 = wb['SKU Table']
for row in ws2.iter_rows():
    r = row[0].row
    for cell in row:
        if r == 1:              cell.protection = locked
        elif r in (2, 3):      cell.protection = locked
        elif 4 <= r <= 10:     cell.protection = unlocked   # rows 4-10 editable
        elif r == 11:           cell.protection = locked    # BD與SKU清單 header
        elif r == 12:           cell.protection = locked    # column headers
        else:                   cell.protection = unlocked  # SKU data rows
ws2.protection.sheet = True
ws2.protection.password = ''
ws2.protection.selectLockedCells = False
ws2.protection.selectUnlockedCells = False

ws0 = wb[wb.sheetnames[0]]
for row in ws0.iter_rows():
    for cell in row:
        cell.protection = locked
ws0.protection.sheet = True
ws0.protection.password = ''
ws0.protection.selectLockedCells = False
ws0.protection.selectUnlockedCells = False

try:
    wb.save(LOCK)
    print(f"Protected template saved → {LOCK}")
except PermissionError:
    print(f"⚠️  無法儲存 {LOCK}（檔案被 Excel 佔用），請關閉後重新執行。")

# ── 3. Create 3 mock files ───────────────────────────────────────────────────
# SKU Table layout (after rows 9-10 inserted):
#   Row  4 : Project Code       / Phase
#   Row  5 : Project Name       / PL-PM
#   Row  6 : 測試負責人          / Fan Mode
#   Row  7 : 測試開始日          / 測試完成日
#   Row  8 : 報告版本            / 備註
#   Row  9 : 申請人 Email        / Product Line      ← system reads C9, F9
#   Row 10 : Sub-series          / 專案名稱(系統)     ← system reads C10, F10
#   Row 11 : BD與SKU清單 (merged header)
#   Row 12 : BD / SKU編號 / GPU / Panel / DDR (column headers)
#   Row 13 : SKU-01 data
#   Row 14 : SKU-02 data

mocks = [
    dict(
        fname="G814JV_EVT_RESULT_NEW_260423.xlsx",
        project_code="G814JV", phase="EVT", os_ver="Win11 24H2",
        req_email="yu.sandra@inventec.com",
        prod_line="Strix", sub_series="Intel", proj_name="Intel NVL-HX(G8-Scar)-BD1",
        sku1="SKU-01", sku2="SKU-02",
        bd1="G814JV-0021", bd2="G814JV-0032",
        cpu1="Intel Core Ultra 9 185HX", cpu2="Intel Core Ultra 7 165HX",
        panel1='18" QHD 240Hz', panel2='18" FHD 144Hz',
        ddr1="64GB DDR5-5600", ddr2="32GB DDR5-5200",
        bat=[8.5,7.9, 11.2,10.8, 9.1,8.6, 12.3,11.9],
        bench=[18542,17890, 2156,2089, 12450,11980, 28930,27650, 7823,7541],
        fps=[87,82, 54,51, 112,108, 245,231, 189,178],
    ),
    dict(
        fname="F607BA_PR_RESULT_NEW_260423.xlsx",
        project_code="F607BA", phase="PR", os_ver="Win11 23H2",
        req_email="chen.kevin@inventec.com",
        prod_line="TUF", sub_series="Intel", proj_name='Intel NVL-BD1 (16")',
        sku1="SKU-01", sku2="SKU-02",
        bd1="F607BA-0011", bd2="F607BA-0022",
        cpu1="Intel Core Ultra 5 125H", cpu2="Intel Core Ultra 5 125H",
        panel1='16" FHD 144Hz', panel2='16" FHD 60Hz',
        ddr1="16GB DDR5-4800", ddr2="16GB DDR5-4800",
        bat=[10.2,11.5, 13.8,15.1, 11.0,12.3, 14.5,16.2],
        bench=[14230,14156, 1876,1871, 7820,5430, 18920,12340, 6890,6750],
        fps=[62,41, 38,25, 89,65, 178,124, 145,98],
    ),
    dict(
        fname="G634JY_ER2_RESULT_R1_260423.xlsx",
        project_code="G634JY", phase="ER2", os_ver="Win11 24H2",
        req_email="lin.amy@inventec.com",
        prod_line="Strix", sub_series="AMD", proj_name="Gator Range (G6/G8-AMD)-BD1",
        sku1="SKU-01", sku2="SKU-02",
        bd1="G634JY-0041", bd2="G634JY-0052",
        cpu1="AMD Ryzen 9 7945HX", cpu2="AMD Ryzen 7 7745HX",
        panel1='16" QHD 240Hz', panel2='16" FHD 165Hz',
        ddr1="32GB DDR5-4800", ddr2="16GB DDR5-4800",
        bat=[7.8,8.9, 10.5,11.8, 8.3,9.6, 11.2,12.9],
        bench=[21890,17320, 1923,1756, 10230,7890, 24560,18970, 7230,6980],
        fps=[74,58, 46,36, 98,81, 215,168, 167,138],
    ),
]

for m in mocks:
    mwb = openpyxl.load_workbook(SRC)

    # ── Performance Test ──────────────────────────────────────────────────────
    ws = mwb['Performance Test']
    ws.protection.sheet = False
    ws.cell(2, 2, m['project_code'])
    ws.cell(2, 4, m['os_ver'])
    ws.cell(2, 6, m['phase'])
    ws.cell(5, 4, m['sku1']); ws.cell(5, 5, m['sku2'])
    for c in range(6, 12):
        for r in range(5, 10):
            ws.cell(r, c, None)
    ws.cell(6, 4, m['bd1']);    ws.cell(6, 5, m['bd2'])
    ws.cell(7, 4, m['cpu1']);   ws.cell(7, 5, m['cpu2'])
    ws.cell(8, 4, m['panel1']); ws.cell(8, 5, m['panel2'])
    ws.cell(9, 4, m['ddr1']);   ws.cell(9, 5, m['ddr2'])
    for i, (v1, v2) in enumerate(zip(m['bat'][::2], m['bat'][1::2])):
        ws.cell(12+i, 4, v1); ws.cell(12+i, 5, v2)
    for i, (v1, v2) in enumerate(zip(m['bench'][::2], m['bench'][1::2])):
        ws.cell(21+i, 4, v1); ws.cell(21+i, 5, v2)
    for i, (v1, v2) in enumerate(zip(m['fps'][::2], m['fps'][1::2])):
        ws.cell(41+i, 4, v1); ws.cell(41+i, 5, v2)

    # ── SKU Table ─────────────────────────────────────────────────────────────
    ws2 = mwb['SKU Table']
    ws2.protection.sheet = False
    ws2.cell(4, 3, m['project_code']); ws2.cell(4, 6, m['phase'])
    ws2.cell(5, 3, f"{m['project_code']} Gaming Laptop")
    ws2.cell(7, 3, '2026-04-23');      ws2.cell(7, 6, '2026-04-23')
    ws2.cell(8, 3, 'v1.0')
    # Row 9: 申請人 Email (C9) + Product Line (F9)
    ws2.cell(9, 3).value = m['req_email']
    ws2.cell(9, 6).value = m['prod_line']
    # Row 10: Sub-series (C10) + 專案名稱(系統) (F10)
    ws2.cell(10, 3).value = m['sub_series']
    ws2.cell(10, 6).value = m['proj_name']
    # SKU data rows 13-14
    ws2.cell(13, 2, m['bd1']);  ws2.cell(13, 3, m['sku1'])
    ws2.cell(13, 4, m['cpu1']); ws2.cell(13, 5, m['panel1']); ws2.cell(13, 6, m['ddr1'])
    ws2.cell(14, 2, m['bd2']);  ws2.cell(14, 3, m['sku2'])
    ws2.cell(14, 4, m['cpu2']); ws2.cell(14, 5, m['panel2']); ws2.cell(14, 6, m['ddr2'])

    # ── 說明 sheet ────────────────────────────────────────────────────────────
    mwb[mwb.sheetnames[0]].protection.sheet = False

    path = os.path.join(OUT, m['fname'])
    mwb.save(path)
    print(f"Saved mock: {path}")

print("\nAll done.")

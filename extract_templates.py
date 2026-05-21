"""
Excel 模板結構資料提取腳本
讀取 ACC, TPR, SOW Perf, HW info 四個 Excel 檔案，提取結構資料輸出 JSON
"""
import json
import sys
import traceback
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

# ── 路徑設定 ─────────────────────────────────────────────────────────────────
BASE = Path(r"G:\HW Gaming\EE 原始報告資料")
OUT  = Path(r"D:\GL HTML\HW_Gaming_Lab\template_data_raw.json")

FILES = {
    "acc":     BASE / "ACC.xlsx",
    "tpr":     BASE / "TPR.xlsx",
    "sow":     BASE / "SOW Perf.xlsx",
    "hwinfo":  BASE / "HW info.xlsx",
}

# ── Color 序列化 ──────────────────────────────────────────────────────────────
def serialize_color(color):
    """將 openpyxl Color 物件轉成可序列化字典"""
    if color is None:
        return None
    try:
        ctype = color.type
    except Exception:
        return None

    try:
        if ctype == "rgb":
            try:
                rgb_val = color.rgb
                if isinstance(rgb_val, str):
                    return {"rgb": rgb_val}
                return {"rgb": str(rgb_val)}
            except Exception:
                return None
        elif ctype == "theme":
            return {"theme": color.theme, "tint": float(color.tint)}
        elif ctype == "indexed":
            return {"indexed": color.indexed}
        elif ctype == "auto":
            # auto color（跟隨系統預設色），不輸出具體色碼
            return {"auto": True}
        else:
            return None
    except Exception:
        # 最後防線：嘗試直接讀 value
        try:
            return {"rgb": str(color.value)}
        except Exception:
            return None

# ── Font 序列化 ───────────────────────────────────────────────────────────────
def serialize_font(font):
    if font is None:
        return None
    return {
        "name":   font.name,
        "size":   font.size,
        "bold":   font.bold,
        "italic": font.italic,
        "color":  serialize_color(font.color),
    }

# ── Fill 序列化 ───────────────────────────────────────────────────────────────
def serialize_fill(fill):
    if fill is None:
        return None
    if fill.fill_type == "solid":
        return {
            "type":    "solid",
            "fgColor": serialize_color(fill.fgColor),
        }
    return None

# ── Border Side 序列化 ────────────────────────────────────────────────────────
def serialize_side(side):
    if side is None:
        return None
    if not side.style:
        return None
    return {
        "style": side.style,
        "color": serialize_color(side.color),
    }

def serialize_border(border):
    if border is None:
        return None
    result = {
        "top":    serialize_side(border.top),
        "bottom": serialize_side(border.bottom),
        "left":   serialize_side(border.left),
        "right":  serialize_side(border.right),
    }
    # 若全部都是 None 則回傳 None
    if all(v is None for v in result.values()):
        return None
    return result

# ── Alignment 序列化 ──────────────────────────────────────────────────────────
def serialize_alignment(align):
    if align is None:
        return None
    result = {
        "horizontal": align.horizontal,
        "vertical":   align.vertical,
        "wrapText":   align.wrapText,
    }
    if all(v is None for v in result.values()):
        return None
    return result

# ── 主要 Sheet 提取函式 ───────────────────────────────────────────────────────
def extract_sheet(ws, tpr_performance_mode=False, header_only_mode=False):
    """
    從一個 worksheet 提取結構資料。
    tpr_performance_mode: 若 True 套用 TPR Performance 特殊欄位過濾規則
    """
    result = {}

    # ── Merged Cells ──────────────────────────────────────────────────────────
    result["mergedCells"] = [str(rng) for rng in ws.merged_cells.ranges]

    # ── Column Widths ─────────────────────────────────────────────────────────
    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width is not None:
            col_widths[col_letter] = col_dim.width
    result["colWidths"] = col_widths

    # ── Row Heights ───────────────────────────────────────────────────────────
    row_heights = {}
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height is not None:
            row_heights[str(row_idx)] = row_dim.height
    result["rowHeights"] = row_heights

    # ── Freeze Panes ──────────────────────────────────────────────────────────
    result["freezePanes"] = str(ws.freeze_panes) if ws.freeze_panes else None

    # ── Tab Color ─────────────────────────────────────────────────────────────
    try:
        if ws.sheet_properties and ws.sheet_properties.tabColor:
            result["tabColor"] = serialize_color(ws.sheet_properties.tabColor)
        else:
            result["tabColor"] = None
    except Exception:
        result["tabColor"] = None

    # ── Cells ─────────────────────────────────────────────────────────────────
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            r = cell.row
            c = cell.column

            # TPR Performance 特殊過濾規則
            if tpr_performance_mode:
                if c <= 2:
                    pass  # A/B 欄全行
                elif 3 <= c <= 17:
                    if r > 2: continue  # C-Q 欄只留前 2 行
                else:
                    continue  # 其他欄忽略

            # Overall status / Tool Version：只保留 A 欄全行 + 前 2 列
            if header_only_mode:
                if c == 1:
                    pass  # A 欄全行保留
                elif r <= 2:
                    pass  # 前 2 列全欄保留
                else:
                    continue  # 其他略過

            v = cell.value

            # 跳過完全空白且無樣式的 cell
            if v is None and not cell.has_style:
                continue

            # 判斷值類型
            if isinstance(v, str):
                cell_v = v
            elif isinstance(v, bool):
                # bool 在 Python 是 int 的子類，先判 bool
                cell_v = None
            elif isinstance(v, (int, float)):
                cell_v = None
            else:
                # datetime, None, 其他 → 輸出 None（若非字串均清除）
                cell_v = None

            # 如果值已為 None 且無樣式則跳過
            if cell_v is None and not cell.has_style:
                continue

            entry = {
                "r": r,
                "c": c,
                "v": cell_v,
            }

            # ── 樣式提取 ──────────────────────────────────────────────────────
            if cell.has_style:
                font_data   = serialize_font(cell.font)
                fill_data   = serialize_fill(cell.fill)
                border_data = serialize_border(cell.border)
                align_data  = serialize_alignment(cell.alignment)
                numfmt      = cell.number_format

                if font_data:
                    entry["font"] = font_data
                if fill_data:
                    entry["fill"] = fill_data
                if border_data:
                    entry["border"] = border_data
                if align_data:
                    entry["align"] = align_data
                if numfmt and numfmt != "General":
                    entry["numFmt"] = numfmt

            cells.append(entry)

    result["cells"] = cells
    return result

# ── 主程式 ────────────────────────────────────────────────────────────────────
def main():
    output = {}
    stats  = {}

    # ── ACC.xlsx ──────────────────────────────────────────────────────────────
    print("讀取 ACC.xlsx ...")
    wb_acc = openpyxl.load_workbook(FILES["acc"], data_only=True)
    print(f"  ACC Sheets: {wb_acc.sheetnames}")

    ws = wb_acc["Cinebench R23 30-loop"]
    output["acc_cinebench"] = extract_sheet(ws)
    stats["acc_cinebench"] = len(output["acc_cinebench"]["cells"])
    print(f"  acc_cinebench: {stats['acc_cinebench']} cells")

    ws = wb_acc["3Dmark Benchmark"]
    output["acc_3dmark"] = extract_sheet(ws)
    stats["acc_3dmark"] = len(output["acc_3dmark"]["cells"])
    print(f"  acc_3dmark: {stats['acc_3dmark']} cells")

    wb_acc.close()

    # ── TPR.xlsx ──────────────────────────────────────────────────────────────
    print("讀取 TPR.xlsx ...")
    wb_tpr = openpyxl.load_workbook(FILES["tpr"], data_only=True)
    print(f"  TPR Sheets: {wb_tpr.sheetnames}")

    ws = wb_tpr["Overall status"]
    output["tpr_overall"] = extract_sheet(ws, header_only_mode=True)
    stats["tpr_overall"] = len(output["tpr_overall"]["cells"])
    print(f"  tpr_overall: {stats['tpr_overall']} cells")

    ws = wb_tpr["Tool Version"]
    output["tpr_toolversion"] = extract_sheet(ws, header_only_mode=True)
    stats["tpr_toolversion"] = len(output["tpr_toolversion"]["cells"])
    print(f"  tpr_toolversion: {stats['tpr_toolversion']} cells")

    ws = wb_tpr["Battery life"]
    output["tpr_battery"] = extract_sheet(ws)
    stats["tpr_battery"] = len(output["tpr_battery"]["cells"])
    print(f"  tpr_battery: {stats['tpr_battery']} cells")

    ws = wb_tpr["Performance"]
    output["tpr_performance"] = extract_sheet(ws, tpr_performance_mode=True)
    stats["tpr_performance"] = len(output["tpr_performance"]["cells"])
    print(f"  tpr_performance: {stats['tpr_performance']} cells")

    wb_tpr.close()

    # ── SOW Perf.xlsx ─────────────────────────────────────────────────────────
    print("讀取 SOW Perf.xlsx ...")
    wb_sow = openpyxl.load_workbook(FILES["sow"], data_only=True)
    print(f"  SOW Sheets: {wb_sow.sheetnames}")

    target_sow = "RTX5060_X4 Performance test"
    if target_sow in wb_sow.sheetnames:
        ws = wb_sow[target_sow]
        print(f"  使用指定 sheet: {target_sow}")
    else:
        ws = wb_sow.active
        print(f"  指定 sheet 不存在，使用 active sheet: {ws.title}")

    output["sow_performance"] = extract_sheet(ws)
    stats["sow_performance"] = len(output["sow_performance"]["cells"])
    print(f"  sow_performance: {stats['sow_performance']} cells")

    wb_sow.close()

    # ── HW info.xlsx ──────────────────────────────────────────────────────────
    print("讀取 HW info.xlsx ...")
    wb_hw = openpyxl.load_workbook(FILES["hwinfo"], data_only=True)
    print(f"  HW info Sheets: {wb_hw.sheetnames}")

    ws = wb_hw.worksheets[0]
    print(f"  使用 index 0 sheet: {ws.title}")
    output["hwinfo_summary"] = extract_sheet(ws)
    stats["hwinfo_summary"] = len(output["hwinfo_summary"]["cells"])
    print(f"  hwinfo_summary: {stats['hwinfo_summary']} cells")

    wb_hw.close()

    # ── 寫出 JSON ─────────────────────────────────────────────────────────────
    print(f"\n寫出 JSON 至 {OUT} ...")
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    size_bytes = OUT.stat().st_size
    size_mb    = size_bytes / 1024 / 1024

    print("\n=== 完成 ===")
    print(f"輸出路徑: {OUT}")
    print(f"檔案大小: {size_bytes:,} bytes ({size_mb:.2f} MB)")
    print("\n各 Sheet Cell 數量：")
    for key, count in stats.items():
        print(f"  {key:25s}: {count:6d} cells")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[ERROR] {e}")
        traceback.print_exc()
        sys.exit(1)
